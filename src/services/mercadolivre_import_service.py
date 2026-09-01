from datetime import date
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook

from src.calculators import calculate_sale_profit
from src.database import fetch_all, get_connection, now_iso
from src.utils import mes_referencia_from_date

# Regra gerencial definida para o Mercado Livre.
ML_TAX_PERCENT = 9.0
ML_COMMISSION_PERCENT = 25.0
ML_FIXED_FEE = 7.0
ML_FIXED_FEE_BASE = "pedido"
REPORT_TYPE = "mercadolivre_performance"


def _norm(value) -> str:
    return "" if value is None else str(value).strip()


def _key(value) -> str:
    text = unicodedata.normalize("NFKD", _norm(value).lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text)


def _money(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _norm(value).replace("R$", "").replace(" ", "")
    if not text:
        return 0.0
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _int(value) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


# O relatório usado no DRE tem estas colunas. Não usamos aliases genéricos
# para evitar confundir "Quantidade de vendas" com "Unidades vendidas".
COLUMN_ALIASES = {
    "ad_id": ["ID do anúncio", "ID do anuncio"],
    "produto_nome": ["Anúncio", "Anuncio"],
    "variacao_nome": ["Variação", "Variacao"],
    "sku": ["SKU"],
    "status": ["Status atual"],
    "pedidos": ["Quantidade de vendas"],
    "unidades": ["Unidades vendidas"],
    "faturamento": ["Vendas brutas (BRL)"],
}


def _find_col(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [_key(header) for header in headers]
    for alias in aliases:
        key = _key(alias)
        if key in normalized:
            return normalized.index(key)
    return None


def _find_header_and_columns(ws):
    """Localiza a linha dos cabeçalhos e exige o trio financeiro do relatório."""
    for row_number, raw in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), 1):
        headers = [_norm(v) for v in raw]
        cols = {field: _find_col(headers, aliases) for field, aliases in COLUMN_ALIASES.items()}
        required = ("produto_nome", "pedidos", "unidades", "faturamento")
        if all(cols[field] is not None for field in required):
            return row_number, headers, cols
    raise ValueError(
        "A planilha selecionada não parece ser o relatório de vendas do Mercado Livre. "
        "Preciso das colunas: Anúncio, Quantidade de vendas, Unidades vendidas e Vendas brutas (BRL)."
    )


def _report_period(ws) -> tuple[date, date]:
    text = ""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        text += " " + " ".join(_norm(v) for v in row if v is not None)
    match = re.search(r"de (\d{1,2}) de (\w+) de (\d{4}) até (\d{1,2}) de (\w+) de (\d{4})", text, re.I)
    months = {"janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4, "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12}
    if not match:
        return date.today(), date.today()
    try:
        return (
            date(int(match.group(3)), months[match.group(2).lower()], int(match.group(1))),
            date(int(match.group(6)), months[match.group(5).lower()], int(match.group(4))),
        )
    except (KeyError, ValueError):
        return date.today(), date.today()


def preview_mercadolivre(file_path: str) -> dict:
    path = Path(file_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Relatório"] if "Relatório" in wb.sheetnames else wb.active
        header_row, headers, cols = _find_header_and_columns(ws)
        rows = []
        for raw_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(raw_values)
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))
            name = _norm(values[cols["produto_nome"]])
            orders = _int(values[cols["pedidos"]])
            units = _int(values[cols["unidades"]])
            gross = _money(values[cols["faturamento"]])
            if not name or orders <= 0 or units <= 0 or gross <= 0:
                continue
            ad_id = _norm(values[cols["ad_id"]]) if cols["ad_id"] is not None else ""
            variation = _norm(values[cols["variacao_nome"]]) if cols["variacao_nome"] is not None else ""
            sku = _norm(values[cols["sku"]]) if cols["sku"] is not None else ""
            status = _norm(values[cols["status"]]) if cols["status"] is not None else ""
            rows.append({
                "ad_id": ad_id or f"ML-NOME:{_key(name)}",
                "produto_nome": name,
                "variacao_nome": variation or "Sem variação",
                "sku": sku,
                "status": status,
                "pedidos": orders,
                "unidades": units,
                "faturamento": gross,
            })
        start, end = _report_period(ws)
        return {
            "arquivo": path.name,
            "data_inicio": start,
            "data_fim": end,
            "rows": rows,
            "count": len(rows),
            "pedidos": sum(r["pedidos"] for r in rows),
            "faturamento": sum(r["faturamento"] for r in rows),
            "unidades": sum(r["unidades"] for r in rows),
        }
    finally:
        wb.close()


def _ensure_ml_schema(conn) -> None:
    """Migração leve para instalações existentes, sem apagar dados."""
    columns = {row["name"] for row in conn.execute("PRAGMA table_info(vendas_contabilizadas)").fetchall()}
    if "pedidos" not in columns:
        conn.execute("ALTER TABLE vendas_contabilizadas ADD COLUMN pedidos INTEGER NOT NULL DEFAULT 0")


def _upsert_ml_product(conn, row: dict) -> tuple[int, int]:
    parent_key = f"ML:{row['ad_id']}"
    existing = conn.execute("SELECT id FROM produtos_pai WHERE id_item_shopee = ?", (parent_key,)).fetchone()
    if existing:
        parent_id = int(existing["id"])
        conn.execute("UPDATE produtos_pai SET nome = ? WHERE id = ?", (row["produto_nome"], parent_id))
    else:
        cur = conn.execute("INSERT INTO produtos_pai (id_item_shopee, nome, ativo, criado_em) VALUES (?, ?, 1, ?)", (parent_key, row["produto_nome"], now_iso()))
        parent_id = int(cur.lastrowid)

    variation_key = f"ML:{row['ad_id']}:{row['variacao_nome']}"
    existing = conn.execute("SELECT id FROM variacoes WHERE id_variacao_shopee = ?", (variation_key,)).fetchone()
    sku = row["sku"] or f"ML:{row['ad_id']}"
    if existing:
        variation_id = int(existing["id"])
        conn.execute("UPDATE variacoes SET produto_pai_id = ?, nome_variacao = ?, sku = ? WHERE id = ?", (parent_id, row["variacao_nome"], sku, variation_id))
    else:
        cur = conn.execute(
            "INSERT INTO variacoes (produto_pai_id, id_variacao_shopee, nome_variacao, sku, tipo_produto, ativo, criado_em) VALUES (?, ?, ?, ?, 'pronto', 1, ?)",
            (parent_id, variation_key, row["variacao_nome"], sku, now_iso()),
        )
        variation_id = int(cur.lastrowid)
    return parent_id, variation_id


def _current_cost(conn, variation_id: int) -> float | None:
    row = conn.execute(
        "SELECT custo_unitario FROM custos_variacao WHERE variacao_id = ? AND ativo = 1 ORDER BY criado_em DESC, id DESC LIMIT 1",
        (variation_id,),
    ).fetchone()
    return float(row["custo_unitario"]) if row else None


def find_ml_importations_same_period(data_inicio: date, data_fim: date) -> list[dict]:
    return fetch_all(
        "SELECT id, arquivo_nome, data_inicio, data_fim, mes_referencia, criado_em FROM importacoes WHERE tipo_relatorio = ? AND data_inicio = ? AND data_fim = ? AND status = 'confirmada' ORDER BY criado_em DESC",
        (REPORT_TYPE, data_inicio.isoformat(), data_fim.isoformat()),
    )


def save_mercadolivre_importation(file_path: str, data_inicio: date, data_fim: date, replace_same_period: bool = True) -> dict:
    preview = preview_mercadolivre(file_path)
    with get_connection() as conn:
        _ensure_ml_schema(conn)
        if replace_same_period:
            existing = conn.execute(
                "SELECT id FROM importacoes WHERE tipo_relatorio = ? AND data_inicio = ? AND data_fim = ? AND status = 'confirmada'",
                (REPORT_TYPE, data_inicio.isoformat(), data_fim.isoformat()),
            ).fetchall()
            for row in existing:
                conn.execute("DELETE FROM importacoes WHERE id = ?", (row["id"],))

        cur = conn.execute(
            "INSERT INTO importacoes (arquivo_nome, caminho_arquivo, tipo_relatorio, tipo_periodo, data_inicio, data_fim, mes_referencia, status, criado_em) VALUES (?, ?, ?, 'personalizado', ?, ?, ?, 'confirmada', ?)",
            (Path(file_path).name, str(Path(file_path)), REPORT_TYPE, data_inicio.isoformat(), data_fim.isoformat(), mes_referencia_from_date(data_inicio), now_iso()),
        )
        importacao_id = int(cur.lastrowid)
        inserted = 0
        incomplete = 0
        for row in preview["rows"]:
            parent_id, variation_id = _upsert_ml_product(conn, row)
            cost = _current_cost(conn, variation_id)
            calc = calculate_sale_profit(
                row["faturamento"],
                row["unidades"],
                ML_TAX_PERCENT,
                ML_COMMISSION_PERCENT,
                ML_FIXED_FEE,
                cost,
                quantidade_taxa_fixa=row["pedidos"],
            )
            incomplete += 1 if calc.lucro_incompleto else 0
            conn.execute(
                "INSERT INTO vendas_contabilizadas (importacao_id, produto_pai_id, variacao_id, data_inicio, data_fim, mes_referencia, unidades, pedidos, faturamento, imposto_percentual, comissao_percentual, taxa_fixa_unitaria, imposto_valor, comissao_valor, taxa_fixa_valor, custo_unitario_usado, custo_total, lucro, lucro_incompleto, criado_em) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    importacao_id, parent_id, variation_id, data_inicio.isoformat(), data_fim.isoformat(), mes_referencia_from_date(data_inicio),
                    calc.unidades, row["pedidos"], calc.faturamento, calc.imposto_percentual, calc.comissao_percentual, calc.taxa_fixa_unitaria,
                    calc.imposto_valor, calc.comissao_valor, calc.taxa_fixa_valor, calc.custo_unitario, calc.custo_total, calc.lucro,
                    1 if calc.lucro_incompleto else 0, now_iso(),
                ),
            )
            inserted += 1
    return {
        "importacao_id": importacao_id,
        "inserted": inserted,
        "incomplete": incomplete,
        "pedidos": preview["pedidos"],
        "faturamento": preview["faturamento"],
        "unidades": preview["unidades"],
        "data_inicio": preview["data_inicio"],
        "data_fim": preview["data_fim"],
    }
